"""
MIRAVERSEOSX SQLite Importer
Requirements: pip install openpyxl
Usage: python miraverse_importer.py
Output: miraverse.db (ready to embed in game backend)
"""
import sqlite3, re, os, sys

# Windows consoles default to cp1252, which can't encode the status output below.
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    import openpyxl  # type: ignore[import-not-found]
except ImportError as exc:
    raise SystemExit('openpyxl is required: pip install openpyxl') from exc

XLSX_PATH = 'MIRAVERSEOSX Game Design Document.xlsx'  # update path if needed
DB_PATH   = 'miraverse.db'

FK_SENTINELS = {'NONE', 'ALL', 'N/A', 'UNKNOWN', '', 'NPC_ASH'}

FK_MAP = {
    'NPCs': {
        'Faction_ID': ('Factions', 'Faction_ID'),
        'Region_ID':  ('Regions',  'Region_ID'),
        'House_ID':   ('Houses',   'House_ID'),
    },
    'Houses': {
        'Region_ID':       ('Regions', 'Region_ID'),
        'Rival_House_ID':  ('Houses',  'House_ID'),
        'Allied_House_ID': ('Houses',  'House_ID'),
    },
    'Factions': {
        'Leader_NPC_ID': ('NPCs',    'NPC_ID'),
        'HQ_Region_ID':  ('Regions', 'Region_ID'),
    },
    'Regions': {
        'Controlling_Faction_ID': ('Factions', 'Faction_ID'),
        'Primary_House_ID':       ('Houses',   'House_ID'),
    },
    'Careers': {
        'Starting_Region_ID':  ('Regions',  'Region_ID'),
        'Starting_Faction_ID': ('Factions', 'Faction_ID'),
    },
    'Apps': {
        'Developer_Faction_ID': ('Factions', 'Faction_ID'),
    },
    'Lore_Entries': {
        'Region_ID':  ('Regions',  'Region_ID'),
        'Faction_ID': ('Factions', 'Faction_ID'),
    },
    'Events': {
        'Region_ID':  ('Regions',  'Region_ID'),
        'Faction_ID': ('Factions', 'Faction_ID'),
    },
}

ALL_FK_COLS = set(col for cols in FK_MAP.values() for col in cols)
TABLE_ORDER = ['Regions','Houses','Factions','NPCs','Careers',
               'Modules','Apps','Lore_Entries','Events','Dashboard']

def is_ref_col(col_name):
    """Reference columns hold IDs, so 'NONE'/'N/A' there means NULL. Prose columns
    (Role_Class, Era, Primary_Resource) legitimately contain 'Unknown' as content."""
    if not col_name: return False
    return col_name in ALL_FK_COLS or col_name.lower().endswith(('_id', '_ids'))

def sanitize_col(name):
    if not name: return '_col'
    return re.sub(r'[^a-zA-Z0-9_]', '_', str(name).strip())

def infer_type(col_name, samples):
    col = col_name.lower() if col_name else ''
    int_hints  = {'level','count','approx','tier','limit','hours','players',
                  'questline_count','word_count_approx','prestige_level',
                  'unlock_level','duration_hours','max_players',
                  'min_player_level','stack_limit','danger_level'}
    real_hints = {'value','cost','scale','subscription_cost_per_day','market_value'}
    for h in int_hints:
        if h in col: return 'INTEGER'
    for h in real_hints:
        if h in col: return 'REAL'
    for v in samples:
        if v is None: continue
        if isinstance(v, (bool, int)): return 'INTEGER'
        if isinstance(v, float): return 'REAL'
        if isinstance(v, str):
            s = v.strip()
            try: int(s);   return 'INTEGER'
            except ValueError: pass
            try: float(s); return 'REAL'
            except ValueError: pass
        break
    return 'TEXT'

def coerce(value, dtype, col_name=None, is_pk=False):
    # Never null a primary key -- a bad PK should surface, not vanish.
    if is_ref_col(col_name) and not is_pk:
        if value is None or str(value).strip().upper() in FK_SENTINELS:
            return None
    if value is None: return None
    if dtype == 'INTEGER':
        try: return int(str(value).strip())
        except: return None
    if dtype == 'REAL':
        try: return float(str(value).strip())
        except: return None
    return str(value) if not isinstance(value, str) else value

# ── Parse workbook ────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
sheet_data = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    if sheet_name == 'Dashboard':
        # Walk each row left-to-right pairing key->value. Consuming a cell as a value
        # stops it being re-read as the next key (which produced ('20','') junk rows).
        kv, seen_keys = [], {}
        for r in all_rows:
            c = 0
            while c < len(r) - 1:
                key, val = r[c], r[c+1]
                if key is None or str(key).strip() == '' or val is None or str(val).strip() == '':
                    c += 1
                    continue
                key = str(key).strip()
                if key in seen_keys:            # Key is PRIMARY KEY; disambiguate
                    seen_keys[key] += 1         # instead of losing the row silently
                    key = f'{key} ({seen_keys[key]})'
                else:
                    seen_keys[key] = 0
                kv.append((key, str(val).strip()))
                c += 2
        sheet_data['Dashboard'] = {'headers':['Key','Value'],'types':['TEXT','TEXT'],'rows':kv}
        continue
    if len(all_rows) < 2: continue
    raw_headers, data_rows = all_rows[1], all_rows[2:]
    headers, seen = [], {}
    for h in raw_headers:
        col = sanitize_col(h) if h else f'_col{len(headers)}'
        if col in seen: seen[col] += 1; col = f'{col}_{seen[col]}'
        else: seen[col] = 0
        headers.append(col)
    samples = [[] for _ in headers]
    for row in data_rows:
        for i, cell in enumerate(row):
            if i < len(headers) and cell is not None: samples[i].append(cell)
    col_types = [infer_type(headers[i], samples[i]) for i in range(len(headers))]
    parsed_rows = []
    for row in data_rows:
        if all(c is None for c in row): continue
        padded = list(row) + [None]*(len(headers)-len(row))
        parsed_rows.append(tuple(coerce(padded[i], col_types[i], headers[i], i == 0)
                                 for i in range(len(headers))))
    sheet_data[sheet_name] = {'headers':headers,'types':col_types,'rows':parsed_rows}

# ── Build database ────────────────────────────────────────────────────────────
if os.path.exists(DB_PATH): os.remove(DB_PATH)
con = sqlite3.connect(DB_PATH)
con.execute('PRAGMA journal_mode = WAL')
con.execute('PRAGMA foreign_keys = OFF')
cur = con.cursor()

ordered = TABLE_ORDER + [s for s in sheet_data if s not in TABLE_ORDER]
for tbl in ordered:
    if tbl not in sheet_data: continue
    info = sheet_data[tbl]
    hdrs, types = info['headers'], info['types']
    pk = hdrs[0]
    col_defs = [f'  "{h}" {t} PRIMARY KEY' if h == pk else f'  "{h}" {t}'
                for h, t in zip(hdrs, types)]
    fk_clauses = [f'  FOREIGN KEY ("{col}") REFERENCES "{ref}"("{rcol}")'
                  for col, (ref, rcol) in FK_MAP.get(tbl, {}).items() if col in hdrs]
    cur.execute(f'CREATE TABLE IF NOT EXISTS "{tbl}" (\n'
                + ',\n'.join(col_defs + fk_clauses) + '\n);')
con.commit()

total = 0
for tbl in ordered:
    if tbl not in sheet_data: continue
    info = sheet_data[tbl]
    hdrs, rows = info['headers'], info['rows']
    quoted_headers = ', '.join(f'"{h}"' for h in hdrs)
    sql = (f'INSERT OR IGNORE INTO "{tbl}" ({quoted_headers}) '
          f'VALUES ({", ".join(["?"] * len(hdrs))})')
    for row in rows:
        try: cur.execute(sql, row); total += 1
        except Exception as e: print(f'WARN [{tbl}]: {e}')

# Indexes
for idx, tbl, col in [
    ('idx_npcs_faction','NPCs','Faction_ID'),('idx_npcs_region','NPCs','Region_ID'),
    ('idx_npcs_house','NPCs','House_ID'),('idx_factions_region','Factions','HQ_Region_ID'),
    ('idx_careers_region','Careers','Starting_Region_ID'),
    ('idx_careers_faction','Careers','Starting_Faction_ID'),
    ('idx_events_region','Events','Region_ID'),('idx_events_faction','Events','Faction_ID'),
    ('idx_lore_region','Lore_Entries','Region_ID'),('idx_lore_faction','Lore_Entries','Faction_ID'),
    ('idx_apps_faction','Apps','Developer_Faction_ID'),('idx_modules_tier','Modules','Tier'),
]:
    cur.execute(f'CREATE INDEX IF NOT EXISTS {idx} ON "{tbl}"("{col}")')

con.commit()
con.execute('PRAGMA foreign_keys = ON')
violations = con.execute('PRAGMA foreign_key_check').fetchall()

# The workbook's Dashboard sheet is empty, so derive build metrics from what was
# actually imported rather than leaving the table with zero rows.
if not cur.execute('SELECT COUNT(*) FROM "Dashboard"').fetchone()[0]:
    content_tables = [t for t in ordered if t in sheet_data and t != 'Dashboard']
    metrics = [('Total Rows Imported', str(total)),
               ('Content Tables', str(len(content_tables))),
               ('FK Violations', str(len(violations)))]
    for tbl in content_tables:
        n = cur.execute(f'SELECT COUNT(*) FROM "{tbl}"').fetchone()[0]
        metrics.append((f'{tbl} Count', str(n)))
    cur.executemany('INSERT OR REPLACE INTO "Dashboard" ("Key","Value") VALUES (?,?)', metrics)
    print(f'   Dashboard sheet empty — generated {len(metrics)} derived metrics')

con.commit()
con.close()

print(f"✅ miraverse.db ready — {total} rows, {len(violations)} FK violations")
print(f"   Size: {os.path.getsize(DB_PATH):,} bytes")
